from werkzeug.datastructures import FileStorage
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from web_errors import WebError
from dataclasses import dataclass
from typing import Iterable
from clinics.data_access import get_most_common_clinic
from patients.data_access import patient_from_key_data, add_patient
from patients.patient import Patient
from visits.data_access import first_visit_by_patient_and_date, add_visit
from visits.visit import Visit
from events.data_access import clear_all_events, add_event
from events.event import Event
import uuid
from language_strings.language_string import LanguageString
from util import as_string
from datetime import date, timedelta, datetime
import itertools
import json
from config import DEFAULT_PROVIDER_ID_FOR_IMPORT
import pandas as pd
import dateutil

COLUMNS = [
    'camp',
    'visit_date',
    'visit_type',
    'first_name',
    'surname',
    'number',
    'date_of_birth',
    'age',
    'gender',
    'hometown',
    'home_country',
    'phone',
    'number',
    'doctor',
    # medical history
    'allergies',
    'tobacco',
    'alcohol',
    'drugs',
    'surgeryHx',
    'chronicConditions',
    'currentMedications',
    'vaccinations',
    'cancer',
    'cancerDetails',
    'epilepsy',
    'epilepsyDetails',
    'heartDisease',
    'heartDiseaseDetails',
    'hypertension',
    'hypertensionDetails',
    'thyroidConditions',
    'thyroidConditionsDetails',
    'tuberculosis',
    'tuberculosisDetails',
    'diabetes',
    'diabetesDetails',
    'multipleBirths',
    'multipleBirthsDetails',
    'breechBirths',
    'breechBirthsDetails',
    'domesticViolence',
    'domesticViolenceDetails',
    'alcoholismDrugAddiction',
    'alcoholismDrugAddictionDetails',
    'codependency',
    'codependencyDetails',
    'nutritionalDisorder',
    'nutritionalDisorderDetails',
    'abuse',
    'abuseDetails',
    'sexualAbuse',
    'sexualAbuseDetails',
    'mentalDisorders',
    'mentalDisordersDetails',
    
    # complaint
    'complaint',
    # vitals
    'heart_rate',
    'blood_pressure',
    'sats',
    'temp',
    'respiratory_rate',
    'weight',
    'blood_glucose',
    
    # examination
    # "trauma_physical",
    # "trauma_sexual",
    # "trauma_abuse",
    # "trauma_killing",
    # "trauma_p_seperation",
    # "trauma_threats",
    # "trauma_t_j",
    # "trauma_others",
    # Medicines
    "medicationM1",
    "typeM1",
    "routeM1",
    "dosageM1",
    "daysM1",
    "medicationM2",
    "typeM2",
    "routeM2",
    "dosageM2",
    "daysM2",
    "medicationM3",
    "typeM3",
    "routeM3",
    "dosageM3",
    "daysM3",
    "medicationM4",
    "typeM4",
    "routeM4",
    "dosageM4",
    "daysM4",
    "medicationM5",
    "typeM5",
    "routeM5",
    "dosageM5",
    "daysM5",
    "medicationM6",
    "typeM6",
    "routeM6",
    "dosageM6",
    "daysM6",
    "medicationM7",
    "typeM7",
    "routeM7",
    "dosageM7",
    "daysM7",
    "medicationM8",
    "typeM8",
    "routeM8",
    "dosageM8",
    "daysM8",
    "medicationM9",
    "typeM9",
    "routeM9",
    "dosageM9",
    "daysM9",
    "medicationM10",
    "typeM10",
    "routeM10",
    "dosageM10",
    "daysM10",
    # physiotherapy
    "previousTreatment",
    "previousTreatmentText",
    "complaint",
    "hpi",
    "appearance",
    "eent",
    "heartPulses",
    "lungs",
    "abdomen",
    "genitales",
    "extremities",
    "skin",
    "neuro",
    "isPregnant",
    "lastPeriodFirstDay",
    "assessmentPlan",
    "notes",
    "referral",
    "referralText",
    # Common Problems
    "difficulty_eating",
    "suicidal_ideation",
    "difficulty_sleeping",
    "hours_slept",
    "sleeping_night",
    "sleeping_falling",
    "sleeping_walking",
    "restless",
    "difficulty_stopping_worrying",
    "body_aches",
    "body_aches_where",
    "low_energy",
    "no_interest",
    "guilt",
    "hopeless",
    "flashbacks",
    "hypervigilance",
    "sad_irritable",
    "hallucinations",
    "potential_diagnosis",
    "other_symptom",
    "bed_wetting",
    "defiant",
    "separation_anxiety",
    "communication_difficulties",
    "traumaPhysical",
    "trauma_physical",
    "trauma_sexual",
    "trauma_abuse",
    "trauma_killing",
    "trauma_p_seperation",
    "trauma_threats",
    "trauma_t_j",
    "trauma_others",
    # Inerventions
    "resource_connection",
    "resource_connection_specify",
    "active_listening",
    "psychoeducation",
    "sleep_hygiene",
    "safe_space_imagine",
    "muscle_relaxation",
    "behavioral_activation",
    "grief_letter",
    "changing_thoughts",
    "senses",
    "distraction_techniques",
    "diary_yourself",
    "chair_technique",
    "build_strengths",
    "motivational_interviewing",
    "diaphragmatic_breathing",
    "safety_plan",
    "interpersonal_communication",
    "prayer",
    "supportive_people",
    "other_strategies",
    "psychiatric_medications",
    # Mental Health
    "in_person",
    "tele",
    "female_only",
    "go_clinic",
    "mental_other",
    "acupuncture",
    "send_audio",
    "psychiatry",
    "refill",
    "taking_regularly",
    "concerns_followup",
    "last_follow_up_date",
    "feeling_nervous",
    "no_control_worrying",
    "little_interest",
    "feeling_down",
    "cant_sleep",
    "dont_feel_safe_living",
    "how_you_feel",
    "child_body_feeling",
    "child_away_from_people",
    "child_feeling_happy",
    "child_trouble_sleeping",
    "child_hard_pay_attention",
    "child_feels_alone",
    "body_feelings",
    "away_from_people",
    "feeling_happy",
    "trouble_sleeping",
    "hard_pay_attention",
    "feel_alone",
    "past_week_stress"
    # Nursing Notes
    "bodyHabitus",
    "procedures",
    "observations",
    # Lab
    "dateBt1",
    "testBt1",
    "resultBt1",
    "notesBt1",
    "dateBt2",
    "testBt2",
    "resultBt2",
    "notesBt2",
    "dateBt3",
    "testBt3",
    "resultBt3",
    "notesBt3",
    "dateBt4",
    "testBt4",
    "resultBt4",
    "notesBt4",
    "dateBt5",
    "testBt5",
    "resultBt5",
    "notesBt5",
    "dateI1",
    "testI1",
    "resultI1",
    "notesI1",
    "dateI2",
    "testI2",
    "resultI2",
    "notesI2",
    "dateI3",
    "testI3",
    "resultI3",
    "notesI3",
    # Dental Treatment
    "dental_treatment"
]

@dataclass
class PatientDataRow:
    camp: str = None
    visit_date: str = None
    visit_type: str = None
    first_name: str = None
    surname: str = None
    date_of_birth: str = None
    age: str = None
    gender: str = None
    hometown: str = None
    home_country: str = None
    phone: str = None
    number: str = None
    doctor: str = None
    # Medical History
    allergies: str = None
    tobacco: str = None
    alcohol: str = None
    drugs: str = None
    surgeryHx: str = None
    chronicConditions: str = None
    currentMedications: str = None
    vaccinations: str = None
    cancer: str = None
    cancerDetails: str = None
    epilepsy: str = None
    epilepsyDetails: str = None
    heartDisease: str = None
    heartDiseaseDetails: str = None
    hypertension: str = None
    hypertensionDetails: str = None
    thyroidConditions: str = None
    thyroidConditionsDetails: str = None
    tuberculosis: str = None
    tuberculosisDetails: str = None
    diabetes: str = None
    diabetesDetails: str = None
    multipleBirths: str = None
    multipleBirthsDetails: str = None
    breechBirths: str = None
    breechBirthsDetails: str = None
    domesticViolence: str = None
    domesticViolenceDetails: str = None
    alcoholismDrugAddiction: str = None
    alcoholismDrugAddictionDetails: str = None
    codependency: str = None
    codependencyDetails: str = None
    nutritionalDisorder: str = None
    nutritionalDisorderDetails: str = None
    abuse: str = None
    abuseDetails: str = None
    sexualAbuse: str = None
    sexualAbuseDetails: str = None
    mentalDisorders: str = None
    mentalDisordersDetails: str = None
    # Complaint
    complaint: str = None
    # Vitals
    heart_rate: str = None
    blood_pressure: str = None
    sats: str = None
    temp: str = None
    respiratory_rate: str = None
    weight: str = None
    blood_glucose: str = None
    # Examination
    # trauma_physical: str = None
    # trauma_sexual: str = None
    # trauma_abuse: str = None
    # trauma_killing: str = None
    # trauma_p_seperation: str = None
    # trauma_threats: str = None
    # trauma_t_j: str = None
    # trauma_others: str = None
    # Medicines
    medicationM1: str = None
    typeM1: str = None
    routeM1: str = None
    dosageM1: str = None
    daysM1: str = None
    medicationM2: str = None
    typeM2: str = None
    routeM2: str = None
    dosageM2: str = None
    daysM2: str = None
    medicationM3: str = None
    typeM3: str = None
    routeM3: str = None
    dosageM3: str = None
    daysM3: str = None
    medicationM4: str = None
    typeM4: str = None
    routeM4: str = None
    dosageM4: str = None
    daysM4: str = None
    medicationM5: str = None
    typeM5: str = None
    routeM5: str = None
    dosageM5: str = None
    daysM5: str = None
    medicationM6: str = None
    typeM6: str = None
    routeM6: str = None
    dosageM6: str = None
    daysM6: str = None
    medicationM7: str = None
    typeM7: str = None
    routeM7: str = None
    dosageM7: str = None
    daysM7: str = None
    medicationM8: str = None
    typeM8: str = None
    routeM8: str = None
    dosageM8: str = None
    daysM8: str = None
    medicationM9: str = None
    typeM9: str = None
    routeM9: str = None
    dosageM9: str = None
    daysM9: str = None
    medicationM10: str = None
    typeM10: str = None
    routeM10: str = None
    dosageM10: str = None
    daysM10: str = None
    # Physical Exam
    previousTreatment: str = None
    previousTreatmentText: str = None
    complaint: str = None
    hpi: str = None
    appearance: str = None
    eent: str = None
    heartPulses: str = None
    lungs: str = None
    abdomen: str = None
    genitales: str = None
    extremities: str = None
    skin: str = None
    neuro: str = None
    isPregnant: str = None
    lastPeriodFirstDay: str = None
    assessmentPlan: str = None
    notes: str = None
    referral: str = None
    referralText: str = None
    
    # Common Problems
    difficulty_eating: str = None
    suicidal_ideation: str = None
    difficulty_sleeping: str = None
    hours_slept: str = None
    sleeping_night: str = None
    sleeping_falling: str = None
    sleeping_walking: str = None
    restless: str = None
    difficulty_stopping_worrying: str = None
    body_aches: str = None
    body_aches_where: str = None
    low_energy: str = None
    no_interest: str = None
    guilt: str = None
    hopeless: str = None
    flashbacks: str = None
    hypervigilance: str = None
    sad_irritable: str = None
    hallucinations: str = None
    potential_diagnosis: str = None
    other_symptom: str = None
    bed_wetting: str = None
    defiant: str = None
    separation_anxiety: str = None
    communication_difficulties: str = None
    trauma_physical: str = None
    trauma_sexual: str = None
    trauma_abuse: str = None
    trauma_killing: str = None
    trauma_p_seperation: str = None
    trauma_threats: str = None
    trauma_t_j: str = None
    trauma_others: str = None

    # Inerventions
    resource_connection: str = None
    resource_connection_specify: str = None
    active_listening: str = None
    psychoeducation: str = None
    sleep_hygiene: str = None
    safe_space_imagine: str = None
    muscle_relaxation: str = None
    behavioral_activation: str = None
    grief_letter: str = None
    changing_thoughts: str = None
    senses: str = None
    distraction_techniques: str = None
    diary_yourself: str = None
    chair_technique: str = None
    build_strengths: str = None
    motivational_interviewing: str = None
    diaphragmatic_breathing: str = None
    safety_plan: str = None
    interpersonal_communication: str = None
    prayer: str = None
    supportive_people: str = None
    other_strategies: str = None
    psychiatric_medications: str = None

    # Mental Health
    in_person: str = None
    tele: str = None
    female_only: str = None
    go_clinic: str = None
    mental_other: str = None
    acupuncture: str = None
    send_audio: str = None
    psychiatry: str = None
    refill: str = None
    taking_regularly: str = None
    concerns_followup: str = None
    last_follow_up_date: str = None
    feeling_nervous: str = None
    no_control_worrying: str = None
    little_interest: str = None
    feeling_down: str = None
    cant_sleep: str = None
    dont_feel_safe_living: str = None
    how_you_feel: str = None
    child_body_feeling: str = None
    child_away_from_people: str = None
    child_feeling_happy: str = None
    child_trouble_sleeping: str = None
    child_hard_pay_attention: str = None
    child_feels_alone: str = None
    body_feelings: str = None
    away_from_people: str = None
    feeling_happy: str = None
    trouble_sleeping: str = None
    hard_pay_attention: str = None
    feel_alone: str = None
    past_week_stress: str = None

    # Nursing Notes     
    bodyHabitus: str = None
    procedures: str = None
    observations: str = None

    # Lab
    dateBt1: str = None
    testBt1: str = None
    resultBt1: str = None
    notesBt1: str = None
    dateBt2: str = None
    testBt2: str = None
    resultBt2: str = None
    notesBt2: str = None
    dateBt3: str = None
    testBt3: str = None
    resultBt3: str = None
    notesBt3: str = None
    dateBt4: str = None
    testBt4: str = None
    resultBt4: str = None
    notesBt4: str = None
    dateBt5: str = None
    testBt5: str = None
    resultBt5: str = None
    notesBt5: str = None
    dateI1: str = None
    testI1: str = None
    resultI1: str = None
    notesI1: str = None
    dateI2: str = None
    testI2: str = None
    resultI2: str = None
    notesI2: str = None
    dateI3: str = None
    testI3: str = None
    resultI3: str = None
    notesI3: str = None

    # Dental Treatment
    dental_treatment: str = None


# COLUMN_TYPES = [str, None, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, float, str,
#                 float, float, float, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str,
#                 str, str]


# class PatientDataImporter:
#     def __init__(self, data_file: FileStorage):
#         self.data_filename = self._write_file_to_tempfile(data_file)

#     def run(self):
#         all_rows = [self._parse_row(row) for row in self.iter_data_rows()]
#         print('Creating patients...')
#         self._create_patients(all_rows)
#         print('Creating visits...')
#         self._create_visits(all_rows)

#     def _parse_row(self, row):
#         if len(row) != 41:
#             raise WebError('All data rows must have exactly 41 data points.', 400)
#         values = [self._parse_cell(value, formatter) for value, formatter in zip(row, COLUMN_TYPES)]
#         return PatientDataRow(**dict(zip(COLUMNS, values)))

#     def _parse_cell(self, cell, formatter):
#         if cell == 'Nil' or cell is None:
#             return None
#         if formatter is None:
#             return cell
#         return formatter(cell)

#     @staticmethod
#     def _write_file_to_tempfile(data_file: FileStorage):
#         handle = NamedTemporaryFile('wb', delete=False, suffix='.xlsx')
#         data_file.save(handle)
#         handle.close()
#         print('Upload written to', handle.name)
#         return handle.name

#     def iter_data_rows(self):
#         wb = load_workbook(self.data_filename)
#         ws = wb.active
#         for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=41, values_only=True)):
#             if all(x is None for x in row):
#                 continue
#             yield row

#     def _create_patients(self, rows: Iterable[PatientDataRow]):
#         for patient_data in set(map(lambda r: (r.first_name, r.surname, r.gender, r.home_country, r.age), rows)):
#             first_name, surname, gender, home_country, age = patient_data
#             if not patient_from_key_data(first_name, surname, home_country, self._parse_sex(gender)):
#                 self._create_patient(first_name, surname, home_country, gender, age)

#     def _create_patient(self, given_name, surname, home_country, sex, age):
#         given_name_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': given_name})
#         surname_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': surname})
#         inferred_dob = self._infer_dob(age)
#         patient = Patient(
#             id=str(uuid.uuid4()),
#             edited_at=datetime.now(),
#             given_name=given_name_ls,
#             surname=surname_ls,
#             date_of_birth=inferred_dob,
#             sex=self._parse_sex(sex),
#             country=LanguageString(id=str(uuid.uuid4()), content_by_language={'en': home_country}),
#             phone=None,
#             hometown=None
#         )
#         add_patient(patient)

#     @staticmethod
#     def _parse_sex(sex_str):
#         if sex_str is None:
#             return None
#         elif 'm' in sex_str.lower():
#             return 'M'
#         elif 'f' in sex_str.lower():
#             return 'F'
#         else:
#             return None

#     def _infer_dob(self, age_string):
#         try:
#             int_prefix = int(''.join(itertools.takewhile(str.isnumeric, age_string)))
#             today = date.today()
#             if 'months' in age_string:
#                 return today - timedelta(days=30 * int_prefix)
#             elif 'weeks' in age_string:
#                 return today - timedelta(weeks=int_prefix)
#             elif 'days' in age_string:
#                 return today - timedelta(days=int_prefix)
#             else:
#                 # Assume years if no unit is specified
#                 return today - timedelta(days=365 * int_prefix)
#         except (ValueError, TypeError):
#             return date(1900, 1, 1)

#     @staticmethod
#     def _parse_date(date_str):
#         if isinstance(date_str, date) or isinstance(date_str, datetime):
#             return date_str
#         try:
#             dt = pd.to_datetime(date_str, dayfirst=True).to_pydatetime()
#             return date(year=dt.year, month=dt.month, day=dt.day)
#         except dateutil.parser._parser.ParserError:
#             return None

#     def _create_visits(self, rows: Iterable[PatientDataRow]):
#         for row in rows:
#             patient_id = patient_from_key_data(row.first_name, row.surname, row.home_country, self._parse_sex(row.gender))
#             if not patient_id:
#                 print('Warning: unknown patient; skipping.')
#                 continue
#             visit_date = self._parse_date(row.visit_date)
#             visit_id, visit_timestamp = first_visit_by_patient_and_date(patient_id, visit_date)

#             # TODO: The data import format does not currently specify a clinic. Since
#             # current Hikma instances are single clinic anyway, just get the most common
#             # clinic (in case there is a demo one with few if any visits) and use that.
#             clinic_id = get_most_common_clinic()

#             # TODO: The data import format does not currently specify a provider in a format
#             # that we can use. So for now, use a per-instance default provider that is set via
#             # environment variable.
#             provider_id = DEFAULT_PROVIDER_ID_FOR_IMPORT

#             if visit_id is None:
#                 visit_id = str(uuid.uuid4())
#                 visit_timestamp = datetime.combine(visit_date, datetime.min.time())
#                 visit = Visit(
#                     id=visit_id,
#                     patient_id=patient_id,
#                     edited_at=datetime.now(),
#                     clinic_id=clinic_id,
#                     provider_id=provider_id,
#                     check_in_timestamp=visit_timestamp
#                 )
#                 add_visit(visit)

#                 # Until we implement full deletion, only add visit the first time it is seen.
#                 self._update_events(patient_id, visit_id, visit_timestamp, row)

#     def _update_events(self, patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         # TODO: This will need to be replaced with a mode of deletion that persists through synchronization.
#         # clear_all_events(visit_id)
#         if row.allergies:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Allergies', row.allergies)
#         if any([row.dispensed_medicine_1, row.dispensed_medicine_2,
#                 row.dispensed_medicine_3, row.dispensed_medicine_4]):
#             self._add_dispensed_medicine_event(patient_id, visit_id, visit_timestamp, row)
#         if row.presenting_complaint:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Complaint', row.presenting_complaint)
#         if any([row.heart_rate, row.blood_pressure, row.o2_sats,
#                 row.respiratory_rate, row.temperature, row.blood_glucose]):
#             self._add_vitals_event(patient_id, visit_id, visit_timestamp, row)
#         if row.examination:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Examination', row.examination)
#         if row.diagnosis:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Diagnosis', row.diagnosis)
#         if row.treatment:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Treatment', row.treatment)
#         if row.prescription:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Prescriptions', row.prescription)
#         if row.notes:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Notes', row.notes)
#         if row.camp:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Camp', row.camp)

#     def _add_text_event(self, patient_id: str, visit_id: str, visit_timestamp: datetime,
#                         event_type: str, event_metadata: str):
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type=event_type,
#             event_timestamp=visit_timestamp,
#             event_metadata=event_metadata,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_dispensed_medicine_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         data = [
#             (row.dispensed_medicine_1, row.dispensed_medicine_quantity_1),
#             (row.dispensed_medicine_2, row.dispensed_medicine_quantity_2),
#             (row.dispensed_medicine_3, row.dispensed_medicine_quantity_3),
#             (row.dispensed_medicine_4, row.dispensed_medicine_quantity_4),
#         ]
#         content = '\n'.join([': '.join(r) for r in data if all(r)])
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Medicine Dispensed',
#             event_timestamp=visit_timestamp,
#             event_metadata=content,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_vitals_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         try:
#             diastolic, systolic = row.blood_pressure.split('/')
#         except (ValueError, AttributeError):
#             diastolic = None
#             systolic = None

#         data = {
#             'heartRate': as_string(row.heart_rate),
#             'systolic': as_string(systolic),
#             'diastolic': as_string(diastolic),
#             'sats': as_string(row.o2_sats),
#             'temp': as_string(row.temperature),
#             'respiratoryRate': as_string(row.respiratory_rate),
#             'bloodGlucose': as_string(row.blood_glucose)
#         }

#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Vitals',
#             event_timestamp=visit_timestamp,
#             event_metadata=json.dumps(data),
#             edited_at=datetime.now(),
#         )
#         add_event(event)
